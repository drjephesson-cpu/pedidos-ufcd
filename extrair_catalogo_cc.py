# -*- coding: utf-8 -*-
"""Extrai catálogo do Centro Cirúrgico a partir das 4 planilhas de pedido."""

from __future__ import annotations

import json
import math
import re
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook

NOVA = Path(r"C:\Users\jephesson.santos\OneDrive - EBSERH\Área de Trabalho\Nova pasta")
OUT = Path(__file__).resolve().parent / "data" / "catalogo_cc.json"

# Ordem das categorias da unidade (fluxo de pedido)
ABA_ORDEM = [
    ("med_caf", "Medicamentos CAF", "Pedido Medicamentos para CAF.xlsx"),
    ("med_farm", "Medicamentos Farmácia Central", "Pedido Medicamentos para FARM CENTRAL.xlsx"),
    ("materiais", "Materiais", "Pedido Materiais COM FILTRO.xlsx"),
    ("fios", "Fios de sutura", "Pedido Fios de sutura COM FILTRO.xlsx"),
]

SKIP_SHEET = re.compile(r"estoque|cole aqui", re.I)
CAT_HINT = re.compile(
    r"^(ampolas|soluç|soluc|pomada|loç|loc|reserva|geladeira|termol|medicamentos\s+termo|"
    r"cremes|pós|pos\b|soros|líquid|liquid|almox)",
    re.I,
)
NOISE = re.compile(
    r"gerar|transfer|e-mail|email|legenda|funcion|data:|n[uú]mero|pedido\s*\d|"
    r"@ebserh|pedir\s*=|filtro|solicitar",
    re.I,
)


def normaliza_codigo(val):
    if val is None or val == "":
        return None
    if isinstance(val, str):
        s = val.strip()
        if s.upper() in ("#VALUE!", "#REF!", "#N/A", "NÃO", "NAO", "SIM"):
            return None
        if not re.match(r"^-?\d+(\.0+)?$", s):
            try:
                return str(int(float(s.replace(",", "."))))
            except Exception:
                return None
        try:
            return str(int(float(s)))
        except Exception:
            return None
    if isinstance(val, float) and math.isnan(val):
        return None
    try:
        return str(int(float(val)))
    except Exception:
        return None


def _is_category(val) -> bool:
    if not isinstance(val, str):
        return False
    s = val.strip()
    if len(s) < 3 or NOISE.search(s):
        return False
    if normaliza_codigo(s):
        return False
    if CAT_HINT.search(s) or s.isupper() or (
        s == s.title() and " " not in s and len(s) < 40
    ):
        # headers like AMPOLAS, SOLUÇÕES, POMADAS, RESERVA TERAPÊUTICA
        if any(c.isdigit() for c in s) and not CAT_HINT.search(s):
            return False
        return True
    if CAT_HINT.search(s):
        return True
    # short all-caps / category-like without digits
    letters = re.sub(r"[^A-Za-zÀ-ÿ]", "", s)
    if letters and letters.upper() == letters and len(s) <= 40 and not any(
        ch.isdigit() for ch in s
    ):
        return True
    return False


def _num(val):
    if val is None or val == "":
        return None
    if isinstance(val, str) and val.strip().upper() in ("#VALUE!", "NÃO", "NAO", "SIM"):
        return None
    try:
        return float(val)
    except Exception:
        return None


def _item(codigo, descricao, est_min, grupo=None):
    if not codigo or not descricao:
        return None
    if est_min is None:
        return None
    desc = str(descricao).strip()
    if not desc or NOISE.search(desc) or _is_category(desc):
        return None
    row = {
        "codigo": int(codigo) if codigo.isdigit() else codigo,
        "descricao": desc,
        "estoque_minimo": float(est_min),
        "ponto_pedido": float(est_min),  # pedir quando abaixo do mínimo
        "caixa_com": 1.0,
        "modelo": "minimo",
    }
    if grupo:
        row["grupo"] = grupo
    return row


def parse_single_column(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    header_i = None
    cod_i = desc_i = min_i = None
    for i, row in enumerate(rows[:15]):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        for j, h in enumerate(cells):
            if h in ("aghu", "cod aghu", "cod. aghu", "cód. aghu", "codigo", "código"):
                header_i = i
                cod_i = j
            if header_i == i and h in (
                "fios cirúrgicos padronizados cc",
                "materia",
                "matéria",
                "medicamento",
                "descrição",
                "descricao",
                "nome",
            ):
                desc_i = j
            if header_i == i and (
                "est" in h and "m" in h
            ) and ("mín" in h or "min" in h or "mìn" in h or "mın" in h or "mín" in h or "min." in h or h.startswith("est")):
                if min_i is None:
                    min_i = j
        if header_i == i:
            # fallbacks
            if desc_i is None and cod_i is not None:
                desc_i = cod_i + 1
            if min_i is None and cod_i is not None:
                min_i = cod_i + 2
            break

    if header_i is None or cod_i is None:
        # Materiais: col A empty, header in row 2
        for i, row in enumerate(rows[:10]):
            for j, c in enumerate(row):
                if isinstance(c, str) and "aghu" in c.lower():
                    header_i = i
                    cod_i = j
                    desc_i = j + 1
                    min_i = j + 2
                    break
            if header_i is not None:
                break
    if header_i is None:
        return []

    itens = []
    seen = set()
    for row in rows[header_i + 1 :]:
        if not row or cod_i >= len(row):
            continue
        # skip section titles / footers
        first = row[cod_i]
        if isinstance(first, str) and (
            NOISE.search(first) or first.upper().startswith("PEDIDO")
        ):
            continue
        cod = normaliza_codigo(row[cod_i])
        if not cod:
            continue
        desc = row[desc_i] if desc_i < len(row) else None
        emin = _num(row[min_i] if min_i < len(row) else None)
        it = _item(cod, desc, emin)
        if not it or cod in seen:
            continue
        seen.add(cod)
        itens.append(it)
    return itens


def parse_two_column(ws) -> list[dict]:
    """Lê layout em duas colunas e reordena: por grupo, coluna esquerda depois direita."""
    rows = list(ws.iter_rows(values_only=True))
    header_i = None
    left0 = right0 = None
    for i, row in enumerate(rows[:10]):
        hits = []
        for j, c in enumerate(row):
            if isinstance(c, str) and c.strip().upper() == "AGHU":
                hits.append(j)
        if len(hits) >= 2:
            header_i = i
            left0, right0 = hits[0], hits[1]
            break
        if len(hits) == 1 and header_i is None:
            # pode ser single — caller decides
            pass
    if header_i is None or left0 is None or right0 is None:
        return parse_single_column(ws)

    # por coluna: lista de (grupo, item)
    def scan(col0: int) -> list[tuple[str, dict]]:
        grupo = "Geral"
        out = []
        seen_local = set()
        for row in rows[header_i + 1 :]:
            if not row or col0 >= len(row):
                continue
            cell = row[col0]
            # category header in codigo column
            if isinstance(cell, str) and _is_category(cell):
                grupo = cell.strip()
                continue
            if isinstance(cell, str) and NOISE.search(cell):
                break
            cod = normaliza_codigo(cell)
            if not cod:
                # category might be only text spanning
                if isinstance(cell, str) and cell.strip() and not NOISE.search(cell):
                    if _is_category(cell) or (
                        not any(ch.isdigit() for ch in cell) and len(cell) < 50
                    ):
                        if CAT_HINT.search(cell) or cell.isupper():
                            grupo = cell.strip()
                continue
            desc = row[col0 + 1] if col0 + 1 < len(row) else None
            emin = _num(row[col0 + 2] if col0 + 2 < len(row) else None)
            it = _item(cod, desc, emin, grupo=grupo)
            if not it or cod in seen_local:
                continue
            seen_local.add(cod)
            out.append((grupo, it))
        return out

    left = scan(left0)
    right = scan(right0)

    # Ordem de grupos: aparição na esquerda, depois novos da direita
    order: list[str] = []
    buckets: OrderedDict[str, list] = OrderedDict()

    def add(grupo, it):
        if grupo not in buckets:
            buckets[grupo] = []
            order.append(grupo)
        # evita duplicar código entre colunas
        if any(str(x["codigo"]) == str(it["codigo"]) for x in buckets[grupo]):
            return
        buckets[grupo].append(it)

    for g, it in left:
        add(g, it)
    for g, it in right:
        add(g, it)

    # Dentro do grupo, mantém ordem de aparição (já inserida)
    itens = []
    seen = set()
    for g in order:
        for it in buckets[g]:
            k = str(it["codigo"])
            if k in seen:
                continue
            seen.add(k)
            itens.append(it)
    return itens


def extract_file(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    try:
        for name in wb.sheetnames:
            if SKIP_SHEET.search(name):
                continue
            ws = wb[name]
            # detectar duas colunas
            rows_preview = list(ws.iter_rows(min_row=1, max_row=8, values_only=True))
            aghu_hits = 0
            for row in rows_preview:
                for c in row:
                    if isinstance(c, str) and c.strip().upper() == "AGHU":
                        aghu_hits += 1
            if aghu_hits >= 2:
                itens = parse_two_column(ws)
            else:
                itens = parse_single_column(ws)
            if itens:
                return itens
        return []
    finally:
        wb.close()


def main():
    abas = []
    itens_map = {}
    for aba_id, titulo, fname in ABA_ORDEM:
        path = NOVA / fname
        if not path.exists():
            raise SystemExit(f"Arquivo não encontrado: {path}")
        itens = extract_file(path)
        abas.append({"id": aba_id, "titulo": titulo, "count": len(itens)})
        itens_map[aba_id] = itens
        print(f"{titulo}: {len(itens)} itens")
        if itens:
            print(f"  1º: {itens[0]['codigo']} {itens[0]['descricao'][:50]}")
            print(f"  últ: {itens[-1]['codigo']} {itens[-1]['descricao'][:50]}")
            grupos = [i.get("grupo") for i in itens if i.get("grupo")]
            if grupos:
                # unique preserve order
                seen = []
                for g in grupos:
                    if g not in seen:
                        seen.append(g)
                print(f"  grupos: {seen}")

    catalogo = {
        "unidade": "cc",
        "titulo": "Centro Cirúrgico",
        "modelo": "minimo",
        "abas": abas,
        "itens": itens_map,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(catalogo, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nSalvo em {OUT}")
    print("Total:", sum(a["count"] for a in abas))


if __name__ == "__main__":
    main()
