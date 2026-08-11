# -*- coding: utf-8 -*-
"""Pedido de Estoque — Farmácia UFCD

Replica a lógica da planilha 'teste pedidos':
- Match por Cód. AGHU
- Pedir? = ponto_pedido > estoque
- Quanto Pedir? = CEILING(estoque_minimo - estoque, caixa_com)
  (ou valor manual, se informado)
"""

from __future__ import annotations

import io
import json
import math
import os
import re
import tempfile
from datetime import date, datetime
from pathlib import Path

try:
    from dotenv import load_dotenv

    load_dotenv()
except ImportError:
    pass

from flask import (
    Flask,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from werkzeug.utils import secure_filename

from auth import (
    UserStore,
    admin_required,
    current_user,
    is_admin,
    login_required,
)
from db import (
    criar_pedido,
    excluir_pedido,
    init_db,
    listar_filtros_historico,
    listar_pedidos,
    load_catalogo_parametros,
    load_estoque_db,
    obter_pedido,
    save_catalogo_parametro,
    save_estoque_db,
    using_neon,
)
from pdf_util import gerar_pdf_pedido

BASE = Path(__file__).resolve().parent
DATA = BASE / "data"
# No Vercel o filesystem do deploy é só leitura; gravações vão para /tmp
ON_VERCEL = bool(os.environ.get("VERCEL"))
WRITABLE = Path("/tmp/pedido-estoque") if ON_VERCEL else DATA
WRITABLE.mkdir(parents=True, exist_ok=True)
DATA.mkdir(exist_ok=True)

CATALOGO_PATH = DATA / "catalogo.json"  # UFCD (compat)
ESTOQUE_PATH = WRITABLE / "estoque_atual.json"  # UFCD legado
USERS_DB = WRITABLE / "users.db"
PEDIDOS_SQLITE = WRITABLE / "pedidos.db"

UNIDADES = {
    "ufcd": {
        "id": "ufcd",
        "titulo": "UFCD",
        "catalogo": DATA / "catalogo.json",
        "estoque_file": "estoque_ufcd.json",
        "legacy_estoque": ["estoque_atual.json"],
        "modelo": "ponto",
        "hint_estoque": "EstoqueFarmacia (saldo farmácia central)",
        "mostra_ponto_caixa": True,
        "aba_ordem": None,  # usa ABA_ORDEM_PADRAO
    },
    "cc": {
        "id": "cc",
        "titulo": "Centro Cirúrgico",
        "catalogo": DATA / "catalogo_cc.json",
        "estoque_file": "estoque_cc.json",
        "legacy_estoque": [],
        "modelo": "minimo",
        "hint_estoque": "EstoqueFarmaciaBloco (saldo farmácia do bloco)",
        "mostra_ponto_caixa": False,
        "aba_ordem": [
            "med_caf",
            "med_farm",
            "materiais",
            "fios",
        ],
    },
}


def resolve_unidade(raw: str | None = None, *, persist: bool | None = None) -> str:
    from_request = raw is None
    if from_request:
        raw = (
            request.args.get("unidade")
            or request.form.get("unidade")
            or session.get("unidade")
            or "ufcd"
        )
    u = str(raw).strip().lower()
    if u in ("centro", "centro_cirurgico", "bloco", "cc"):
        u = "cc"
    if u not in UNIDADES:
        u = "ufcd"
    if persist is None:
        persist = from_request
    if persist:
        session["unidade"] = u
    return u


def unidade_cfg(unidade: str | None = None) -> dict:
    if unidade is None:
        uid = resolve_unidade()
    else:
        uid = resolve_unidade(unidade, persist=False)
    return UNIDADES[uid]


app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "pedido-estoque-ufcd-local")
# Mesmo banco dos pedidos (Neon na nuvem; SQLite local). /tmp no Vercel é efêmero.
users = UserStore(None if using_neon() else PEDIDOS_SQLITE)

try:
    init_db(None if using_neon() else PEDIDOS_SQLITE)
except Exception as exc:  # pragma: no cover
    print("aviso init_db:", exc)


@app.context_processor
def inject_globals():
    uid = session.get("unidade") or "ufcd"
    if uid not in UNIDADES:
        uid = "ufcd"
    return {
        "user": current_user(),
        "is_admin": is_admin(),
        "usando_neon": using_neon(),
        "unidades": UNIDADES,
        "unidade_id": uid,
        "unidade": UNIDADES[uid],
    }

# Abas na mesma ordem da planilha (exceto Estoque)
ABA_ORDEM_PADRAO = [
    "COMPRIMIDOS",
    "SOLUÇÕES",
    "Geladeira",
    "PÓS",
    "CREMES e Pomadas",
    "M.P.P.S",
    "SOROS",
    "LÍQUIDOS E GOTAS",
    "AMPOLAS",
    "RESERVA TERPÊUTICA",
    "ALMOXARIFADO",
]


def ceiling_math(value: float, significance: float) -> float:
    """Equivalente ao CEILING.MATH do Excel."""
    if significance in (0, None):
        significance = 1
    if value <= 0:
        return 0
    return math.ceil(value / significance) * significance


def load_catalogo(unidade: str | None = None) -> dict:
    cfg = unidade_cfg(unidade)
    path = cfg["catalogo"]
    if not path.exists():
        # fallback UFCD legado
        if cfg["id"] == "ufcd" and CATALOGO_PATH.exists():
            path = CATALOGO_PATH
        else:
            return {"abas": [], "itens": {}, "modelo": cfg["modelo"]}
    data = json.loads(path.read_text(encoding="utf-8"))
    data.setdefault("modelo", cfg["modelo"])
    return data


def save_catalogo(catalogo: dict, unidade: str | None = None) -> None:
    cfg = unidade_cfg(unidade)
    path = cfg["catalogo"]
    path.write_text(
        json.dumps(catalogo, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def estoque_path_for(unidade: str | None = None) -> Path:
    cfg = unidade_cfg(unidade)
    primary = WRITABLE / cfg["estoque_file"]
    if primary.exists():
        return primary
    for name in cfg.get("legacy_estoque") or []:
        for base in (WRITABLE, DATA):
            p = base / name
            if p.exists():
                return p
    return primary


def _uid_storage(unidade: str | None = None) -> str:
    if unidade:
        return resolve_unidade(unidade, persist=False)
    try:
        from flask import has_request_context

        if has_request_context():
            u = session.get("unidade")
            if u in UNIDADES:
                return u
    except Exception:
        pass
    return "ufcd"


def load_estoque(unidade: str | None = None) -> dict:
    """Payload {saldos, meta} ou legado."""
    uid = _uid_storage(unidade)
    if using_neon():
        return load_estoque_db(uid, None) or {}
    path = estoque_path_for(uid)
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def save_estoque(mapa: dict, meta: dict | None = None, unidade: str | None = None) -> None:
    uid = _uid_storage(unidade)
    payload_meta = meta or {}
    if using_neon():
        save_estoque_db(uid, mapa, payload_meta, None)
        return
    cfg = UNIDADES[uid]
    path = WRITABLE / cfg["estoque_file"]
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {"saldos": mapa, "meta": payload_meta}
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def estoque_saldos(unidade: str | None = None) -> dict:
    raw = load_estoque(unidade)
    if isinstance(raw, dict) and "saldos" in raw:
        return {str(k): v for k, v in raw["saldos"].items()}
    return {str(k): v for k, v in raw.items()}


def estoque_meta(unidade: str | None = None) -> dict:
    raw = load_estoque(unidade)
    if isinstance(raw, dict) and "meta" in raw:
        return raw["meta"]
    return {}


def manuais_da_unidade(unidade: str | None = None) -> dict:
    uid = resolve_unidade(unidade, persist=False) if unidade is not None else resolve_unidade()
    raw = session.get("manuais", {})
    if not isinstance(raw, dict):
        return {}
    # legado: mapa plano codigo→valor (UFCD)
    if raw and not any(k in UNIDADES for k in raw):
        return raw if uid == "ufcd" else {}
    bucket = raw.get(uid) or {}
    return bucket if isinstance(bucket, dict) else {}


def set_manual_unidade(unidade: str, codigo: str, valor) -> None:
    uid = resolve_unidade(unidade, persist=False)
    raw = session.get("manuais", {})
    if not isinstance(raw, dict):
        raw = {}
    if raw and not any(k in UNIDADES for k in raw):
        raw = {"ufcd": dict(raw)}
    bucket = dict(raw.get(uid) or {})
    if valor in (None, ""):
        bucket.pop(codigo, None)
    else:
        bucket[codigo] = valor
    raw[uid] = bucket
    session["manuais"] = raw
    session.modified = True


def ordenar_abas(abas: list, unidade: str | None = None) -> list:
    cfg = unidade_cfg(unidade)
    ordem_ids = cfg.get("aba_ordem") or ABA_ORDEM_PADRAO
    ordem = {n: i for i, n in enumerate(ordem_ids)}
    return sorted(abas, key=lambda a: ordem.get(a["id"], 100 + abas.index(a)))

def normaliza_codigo(val) -> str | None:
    if val is None or val == "":
        return None
    if isinstance(val, str):
        val = val.strip()
        if not val or not re.match(r"^-?\d+(\.0+)?$", val):
            # tenta número puro
            try:
                return str(int(float(val.replace(",", "."))))
            except Exception:
                return None
        try:
            return str(int(float(val)))
        except Exception:
            return None
    try:
        return str(int(float(val)))
    except Exception:
        return None


def parse_estoque_xlsx(file_storage) -> tuple[dict, dict]:
    """Lê EstoqueFarmacia: colunas codigo + saldo_farmacia_central (ou 8ª col)."""
    wb = load_workbook(file_storage, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return {}, {}

    header_l = [str(h).strip().lower() if h is not None else "" for h in header]
    cod_idx = None
    saldo_idx = None
    for i, h in enumerate(header_l):
        if h in ("codigo", "código", "cod_aghu", "cód. aghu", "cod. aghu"):
            cod_idx = i
        if h in (
            "saldo_farmacia_central",
            "saldo_farmacia_bloco",
            "saldo",
            "estoque",
            "estoque_aghu",
            "qtde",
            "quantidade",
        ):
            saldo_idx = i

    # fallback: 1ª coluna = código, 8ª = saldo (como na planilha)
    if cod_idx is None:
        cod_idx = 0
    if saldo_idx is None:
        saldo_idx = 7 if len(header) >= 8 else len(header) - 1

    mapa: dict[str, float | None] = {}
    for row in rows:
        if not row:
            continue
        cod = normaliza_codigo(row[cod_idx] if cod_idx < len(row) else None)
        if not cod:
            continue
        saldo_raw = row[saldo_idx] if saldo_idx < len(row) else None
        if saldo_raw is None or saldo_raw == "":
            mapa[cod] = None  # sem saldo informado
        else:
            try:
                mapa[cod] = float(saldo_raw)
            except Exception:
                mapa[cod] = None

    arquivo = getattr(file_storage, "filename", None) or getattr(
        file_storage, "name", None
    )
    if arquivo:
        arquivo = Path(str(arquivo)).name
    meta = {
        "arquivo": arquivo or "",
        "itens": len(mapa),
        "importado_em": date.today().isoformat(),
    }
    return mapa, meta


def _is_header_cell(val) -> bool:
    if not isinstance(val, str):
        return False
    vu = val.upper().replace("Ó", "O").replace("Ê", "E")
    return "AGHU" in vu or "CODIGO" in vu


def parse_pedido_workbook(wb) -> dict:
    """Extrai abas de pedido de um workbook openpyxl."""
    catalog: dict = {"abas": [], "itens": {}}
    for name in wb.sheetnames:
        if name.upper() == "ESTOQUE":
            continue
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        header_row = None
        for r_idx, row in enumerate(rows[:12]):
            for cell in row[:12]:
                if _is_header_cell(cell):
                    header_row = r_idx
                    break
            if header_row is not None:
                break
        if header_row is None:
            continue
        items = []
        for row in rows[header_row + 1 :]:
            if not row:
                continue
            cod = normaliza_codigo(row[0])
            if not cod:
                continue
            desc = row[1] if len(row) > 1 else ""
            est_min = row[2] if len(row) > 2 else None
            ponto = row[3] if len(row) > 3 else None
            caixa = row[4] if len(row) > 4 else None
            if est_min is None and ponto is None:
                continue
            try:
                items.append(
                    {
                        "codigo": int(cod),
                        "descricao": str(desc).strip() if desc else "",
                        "estoque_minimo": float(est_min or 0),
                        "ponto_pedido": float(ponto or 0),
                        "caixa_com": float(caixa)
                        if caixa not in (None, 0, "")
                        else 1.0,
                    }
                )
            except Exception:
                continue
        catalog["abas"].append({"id": name, "titulo": name, "count": len(items)})
        catalog["itens"][name] = items
    return catalog


def import_pedido_file(file_bytes: bytes, filename: str, password: str | None) -> dict:
    """Importa catálogo de .xlsx/.xls, com suporte a senha (msoffcrypto)."""
    name_lower = (filename or "").lower()
    data = file_bytes

    # Tenta descriptografar se necessário
    try:
        import msoffcrypto

        bio = io.BytesIO(file_bytes)
        office = msoffcrypto.OfficeFile(bio)
        if office.is_encrypted():
            if not password:
                raise ValueError(
                    "Este arquivo está protegido por senha. Informe a senha para importar."
                )
            bio.seek(0)
            office = msoffcrypto.OfficeFile(bio)
            office.load_key(password=password)
            dec = io.BytesIO()
            office.decrypt(dec)
            data = dec.getvalue()
    except ValueError:
        raise
    except Exception as e:
        # se não for encrypted / lib falhar, segue com bytes originais
        if "password" in str(e).lower() or "decrypt" in str(e).lower():
            raise ValueError(f"Não foi possível abrir com a senha informada: {e}") from e

    # Detecta formato
    if data[:2] == b"PK":
        wb = load_workbook(io.BytesIO(data), data_only=True)
        return parse_pedido_workbook(wb)

    # OLE legado: tenta via Excel COM
    if data[:4] == b"\xd0\xcf\x11\xe0":
        return _import_via_excel_com(data, password)

    raise ValueError(
        "Formato não reconhecido. Salve a planilha como .xlsx (Excel) e tente de novo."
    )


def _import_via_excel_com(data: bytes, password: str | None) -> dict:
    try:
        import pythoncom
        import win32com.client
    except ImportError as e:
        raise ValueError(
            "Arquivo legado (.xls criptografado). Salve como .xlsx no Excel ou "
            "instale pywin32 para importação automática."
        ) from e

    pythoncom.CoInitialize()
    tmp = Path(tempfile.gettempdir()) / "pedido_import_tmp.xls"
    tmp.write_bytes(data)
    excel = None
    wb = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        kwargs = {"UpdateLinks": 0, "ReadOnly": True}
        if password:
            wb = excel.Workbooks.Open(str(tmp), Password=password, **kwargs)
        else:
            wb = excel.Workbooks.Open(str(tmp), **kwargs)

        # Salva temporário xlsx e lê com openpyxl
        xlsx_tmp = Path(tempfile.gettempdir()) / "pedido_import_tmp.xlsx"
        wb.SaveAs(str(xlsx_tmp), FileFormat=51)
        wb.Close(False)
        wb = None
        catalog = parse_pedido_workbook(load_workbook(xlsx_tmp, data_only=True))
        try:
            xlsx_tmp.unlink(missing_ok=True)
        except Exception:
            pass
        return catalog
    finally:
        try:
            if wb is not None:
                wb.Close(False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass


def calcular_item(item: dict, saldo, manual=None, modelo: str | None = None) -> dict:
    est_min = float(item["estoque_minimo"])
    ponto = float(item.get("ponto_pedido") if item.get("ponto_pedido") is not None else est_min)
    caixa = float(item.get("caixa_com") or 1)
    if caixa <= 0:
        caixa = 1
    modelo = (modelo or item.get("modelo") or "ponto").lower()

    estoque = None if saldo is None else float(saldo)
    pedir = False
    quanto = None

    if estoque is not None:
        if modelo == "minimo":
            pedir = estoque < est_min
        else:
            pedir = ponto > estoque
        if manual not in (None, ""):
            try:
                quanto = float(manual)
            except Exception:
                quanto = None
        elif pedir:
            if modelo == "minimo":
                quanto = max(0.0, est_min - estoque)
            else:
                quanto = ceiling_math(est_min - estoque, caixa)

    return {
        **item,
        "estoque_aghu": estoque,
        "pedir": pedir,
        "quanto_pedir": quanto,
        "manual": manual if manual not in (None, "") else None,
        "sem_estoque": estoque is None,
        "modelo": modelo,
    }


def montar_aba(
    aba_id: str,
    catalogo: dict,
    saldos: dict,
    manuais: dict,
    modelo: str | None = None,
    parametros: dict | None = None,
) -> list:
    itens = catalogo.get("itens", {}).get(aba_id, [])
    modelo = modelo or catalogo.get("modelo") or "ponto"
    parametros = parametros or {}
    result = []
    for item in itens:
        row = dict(item)
        cod = str(row["codigo"])
        ov = parametros.get(cod)
        if ov:
            if ov.get("estoque_minimo") is not None:
                row["estoque_minimo"] = float(ov["estoque_minimo"])
            if ov.get("ponto_pedido") is not None:
                row["ponto_pedido"] = float(ov["ponto_pedido"])
            if ov.get("caixa_com") is not None:
                row["caixa_com"] = float(ov["caixa_com"])
            row["param_editado"] = True
        saldo = saldos.get(cod)
        if cod not in saldos:
            saldo = None
        manual = manuais.get(cod)
        result.append(calcular_item(row, saldo, manual, modelo=modelo))
    return result


def montar_todos(
    catalogo: dict,
    saldos: dict,
    manuais: dict,
    modelo: str | None = None,
    parametros: dict | None = None,
) -> list:
    """Todos os itens da unidade, com a categoria de origem."""
    out = []
    modelo = modelo or catalogo.get("modelo") or "ponto"
    for aba in catalogo.get("abas", []):
        for it in montar_aba(
            aba["id"], catalogo, saldos, manuais, modelo=modelo, parametros=parametros
        ):
            row = dict(it)
            row["aba"] = aba["titulo"]
            out.append(row)
    return out


def listar_itens_filtrados(
    aba: str | None = None,
    filtro: str = "todos",
    q: str = "",
    unidade: str | None = None,
) -> tuple[list[dict], dict]:
    """Lista itens como na tela (unidade + aba + filtro + busca)."""
    uid = resolve_unidade(unidade)
    catalogo = load_catalogo(uid)
    saldos = estoque_saldos(uid)
    manuais = manuais_da_unidade(uid)
    parametros = load_catalogo_parametros(uid, _db_path())
    abas = catalogo.get("abas") or []
    abas_sorted = ordenar_abas(abas, uid)
    modelo = catalogo.get("modelo") or unidade_cfg(uid)["modelo"]

    aba_atual = aba or "todos"
    filtro = filtro or "todos"
    q = (q or "").strip().lower()

    if aba_atual == "todos":
        itens = montar_todos(
            catalogo, saldos, manuais, modelo=modelo, parametros=parametros
        )
    else:
        itens = montar_aba(
            aba_atual, catalogo, saldos, manuais, modelo=modelo, parametros=parametros
        )
        titulo = next(
            (a["titulo"] for a in abas_sorted if a["id"] == aba_atual), aba_atual
        )
        for it in itens:
            it["aba"] = titulo

    resumo = {"total": len(itens), "pedir": 0, "sem_saldo": 0, "qtd_pedir": 0}
    for it in itens:
        if it["sem_estoque"]:
            resumo["sem_saldo"] += 1
        if it["pedir"]:
            resumo["pedir"] += 1
            if it["quanto_pedir"]:
                resumo["qtd_pedir"] += it["quanto_pedir"]

    if filtro == "pedir":
        itens = [i for i in itens if i["pedir"]]
    elif filtro == "sem":
        itens = [i for i in itens if i["sem_estoque"]]

    if q:
        itens = [
            i
            for i in itens
            if q in str(i["codigo"])
            or q in (i["descricao"] or "").lower()
            or q in (i.get("aba") or "").lower()
            or q in (i.get("grupo") or "").lower()
        ]

    meta = {
        "aba": aba_atual,
        "filtro": filtro,
        "q": q,
        "ver_todos": aba_atual == "todos",
        "abas": abas_sorted,
        "resumo": resumo,
        "unidade": uid,
        "modelo": modelo,
    }
    return itens, meta


def itens_para_pdf(itens: list[dict]) -> list[dict]:
    out = []
    for it in itens:
        out.append(
            {
                "codigo": it.get("codigo"),
                "descricao": it.get("descricao") or "",
                "aba": it.get("aba") or "",
                "estoque_minimo": it.get("estoque_minimo"),
                "ponto_pedido": it.get("ponto_pedido"),
                "caixa_com": it.get("caixa_com"),
                "estoque_aghu": it.get("estoque_aghu"),
                "quantidade": float(it["quanto_pedir"])
                if it.get("quanto_pedir") is not None
                else 0,
                "pedir": bool(it.get("pedir")),
            }
        )
    return out


def _db_path():
    return None if using_neon() else PEDIDOS_SQLITE


def itens_extras_manuais() -> list[dict]:
    return session.get("extras_manuais", [])


def coletar_itens_pedido(incluir_extras: bool = True, unidade: str | None = None) -> list[dict]:
    """Itens com Pedir=Sim (todas as abas) + extras manuais da sessão."""
    uid = resolve_unidade(unidade)
    catalogo = load_catalogo(uid)
    saldos = estoque_saldos(uid)
    manuais = manuais_da_unidade(uid)
    parametros = load_catalogo_parametros(uid, _db_path())
    modelo = catalogo.get("modelo") or unidade_cfg(uid)["modelo"]
    out: list[dict] = []
    vistos: set[str] = set()

    for aba in catalogo.get("abas", []):
        for it in montar_aba(
            aba["id"], catalogo, saldos, manuais, modelo=modelo, parametros=parametros
        ):
            if not it.get("pedir") or not it.get("quanto_pedir"):
                continue
            key = str(it["codigo"])
            vistos.add(key)
            out.append(
                {
                    "codigo": it["codigo"],
                    "descricao": it["descricao"],
                    "aba": aba["titulo"],
                    "quantidade": float(it["quanto_pedir"]),
                    "origem": "manual" if it.get("manual") is not None else "auto",
                    "estoque_aghu": it.get("estoque_aghu"),
                    "estoque_minimo": it.get("estoque_minimo"),
                    "ponto_pedido": it.get("ponto_pedido"),
                    "caixa_com": it.get("caixa_com"),
                    "pedir": True,
                    "unidade": uid,
                }
            )

    if incluir_extras:
        for ex in itens_extras_manuais():
            cod = ex.get("codigo")
            key = str(cod) if cod not in (None, "") else f"manual:{ex.get('descricao')}"
            if key in vistos:
                continue
            out.append(
                {
                    "codigo": int(cod) if cod not in (None, "") else None,
                    "descricao": ex.get("descricao") or "",
                    "aba": ex.get("aba") or "MANUAL",
                    "quantidade": float(ex.get("quantidade") or 0),
                    "origem": "manual",
                    "estoque_aghu": None,
                    "estoque_minimo": None,
                    "ponto_pedido": None,
                    "caixa_com": None,
                    "pedir": True,
                    "unidade": uid,
                }
            )
    return out


def parse_data_pedido(raw: str | None) -> date:
    raw = (raw or "").strip()
    if not raw:
        return date.today()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return date.today()


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("user"):
        return redirect(url_for("index"))
    if request.method == "POST":
        username = request.form.get("username", "")
        password = request.form.get("password", "")
        user = users.authenticate(username, password)
        if user:
            session["user"] = user
            session.pop("manuais", None)
            session.pop("extras_manuais", None)
            nxt = request.args.get("next") or url_for("index")
            if not nxt.startswith("/"):
                nxt = url_for("index")
            flash(f"Olá, {user['nome']}!", "ok")
            return redirect(nxt)
        flash("Usuário ou senha inválidos.", "erro")
    return render_template("login.html")


@app.route("/logout", methods=["POST", "GET"])
def logout():
    session.clear()
    flash("Sessão encerrada.", "ok")
    return redirect(url_for("login"))


@app.route("/usuarios", methods=["GET", "POST"])
@admin_required
def usuarios():
    if request.method == "POST":
        ok, msg = users.create_user(
            username=request.form.get("username", ""),
            password=request.form.get("password", ""),
            role=request.form.get("role", "usuario"),
            nome=request.form.get("nome", ""),
        )
        flash(msg, "ok" if ok else "erro")
        return redirect(url_for("usuarios"))
    return render_template(
        "usuarios.html",
        lista=users.list_users(),
        user=current_user(),
        is_admin=True,
    )


@app.route("/usuarios/<int:user_id>/papel", methods=["POST"])
@admin_required
def usuarios_papel(user_id: int):
    ok, msg = users.set_role(user_id, request.form.get("role", "usuario"))
    flash(msg, "ok" if ok else "erro")
    return redirect(url_for("usuarios"))


@app.route("/usuarios/<int:user_id>/ativo", methods=["POST"])
@admin_required
def usuarios_ativo(user_id: int):
    ativo = request.form.get("ativo") == "1"
    ok, msg = users.set_ativo(user_id, ativo)
    flash(msg, "ok" if ok else "erro")
    return redirect(url_for("usuarios"))


@app.route("/usuarios/<int:user_id>/excluir", methods=["POST"])
@admin_required
def usuarios_excluir(user_id: int):
    ok, msg = users.delete_user(user_id)
    flash(msg, "ok" if ok else "erro")
    return redirect(url_for("usuarios"))


@app.route("/usuarios/<int:user_id>/senha", methods=["POST"])
@admin_required
def usuarios_senha(user_id: int):
    ok, msg = users.change_password(user_id, request.form.get("password", ""))
    flash(msg, "ok" if ok else "erro")
    return redirect(url_for("usuarios"))


@app.route("/")
@login_required
def index():
    uid = resolve_unidade()
    cfg = unidade_cfg(uid)
    catalogo = load_catalogo(uid)
    saldos = estoque_saldos(uid)
    meta = estoque_meta(uid)
    abas = catalogo.get("abas") or []
    abas_sorted = ordenar_abas(abas, uid)

    aba_atual = request.args.get("aba") or "todos"
    filtro = request.args.get("filtro", "todos")
    q = (request.args.get("q") or "").strip().lower()

    manuais = manuais_da_unidade(uid)
    modelo = catalogo.get("modelo") or cfg["modelo"]
    parametros = load_catalogo_parametros(uid, _db_path())
    itens, view_meta = listar_itens_filtrados(aba_atual, filtro, q, unidade=uid)
    resumo = view_meta["resumo"]
    ver_todos = view_meta["ver_todos"]

    contagens = {}
    total_pedir = 0
    for a in abas_sorted:
        lista = montar_aba(
            a["id"], catalogo, saldos, manuais, modelo=modelo, parametros=parametros
        )
        contagens[a["id"]] = sum(1 for i in lista if i["pedir"])
        total_pedir += contagens[a["id"]]
    contagens["todos"] = total_pedir

    # contagens de pedir por unidade (sidebar)
    contagens_unidade = {uid: total_pedir}
    for other_id in UNIDADES:
        if other_id == uid:
            continue
        try:
            other_cat = load_catalogo(other_id)
            other_saldos = estoque_saldos(other_id)
            other_manuais = manuais_da_unidade(other_id)
            other_params = load_catalogo_parametros(other_id, _db_path())
            other_modelo = other_cat.get("modelo") or UNIDADES[other_id]["modelo"]
            n = 0
            for a in other_cat.get("abas") or []:
                n += sum(
                    1
                    for i in montar_aba(
                        a["id"],
                        other_cat,
                        other_saldos,
                        other_manuais,
                        modelo=other_modelo,
                        parametros=other_params,
                    )
                    if i["pedir"]
                )
            contagens_unidade[other_id] = n
        except Exception:
            contagens_unidade[other_id] = 0

    return render_template(
        "index.html",
        abas=abas_sorted,
        aba_atual=aba_atual,
        ver_todos=ver_todos,
        itens=itens,
        resumo=resumo,
        filtro=filtro,
        q=q,
        tem_catalogo=bool(abas),
        tem_estoque=bool(saldos),
        estoque_meta=meta,
        contagens=contagens,
        contagens_unidade=contagens_unidade,
        hoje=date.today().strftime("%d/%m/%Y"),
        total_catalogo=sum(a.get("count", 0) for a in abas_sorted),
        extras=itens_extras_manuais(),
        unidade_id=uid,
        unidade=cfg,
        mostra_ponto_caixa=cfg.get("mostra_ponto_caixa", True),
    )


@app.route("/importar/catalogo", methods=["POST"])
@login_required
def importar_catalogo():
    uid = resolve_unidade()
    f = request.files.get("arquivo")
    password = (request.form.get("senha") or "").strip() or None
    if not f or not f.filename:
        flash("Selecione o arquivo de pedidos (catálogo).", "erro")
        return redirect(url_for("index", unidade=uid))
    try:
        catalog = import_pedido_file(f.read(), f.filename, password)
        if not catalog["abas"]:
            flash("Nenhuma aba de pedido encontrada no arquivo.", "erro")
            return redirect(url_for("index", unidade=uid))
        save_catalogo(catalog, uid)
        total = sum(a["count"] for a in catalog["abas"])
        flash(
            f"Catálogo importado: {len(catalog['abas'])} abas, {total} itens.",
            "ok",
        )
    except Exception as e:
        flash(str(e), "erro")
    return redirect(url_for("index", unidade=uid))


@app.route("/importar/estoque", methods=["POST"])
@login_required
def importar_estoque():
    uid = resolve_unidade()
    cfg = unidade_cfg(uid)
    f = request.files.get("arquivo")
    if not f or not f.filename:
        flash(f"Selecione o arquivo de estoque ({cfg['hint_estoque']}).", "erro")
        return redirect(url_for("index", unidade=uid))
    try:
        mapa, meta = parse_estoque_xlsx(f)
        if not mapa:
            flash("Nenhum item de estoque encontrado.", "erro")
            return redirect(url_for("index", unidade=uid))
        save_estoque(mapa, meta, uid)
        flash(f"Estoque importado: {len(mapa)} códigos ({meta.get('arquivo')}).", "ok")
    except Exception as e:
        flash(f"Erro ao ler estoque: {e}", "erro")
    return redirect(url_for("index", unidade=uid))


@app.route("/api/manual", methods=["POST"])
@login_required
def api_manual():
    data = request.get_json(force=True)
    uid = resolve_unidade(data.get("unidade"))
    cod = str(data.get("codigo", ""))
    valor = data.get("valor")
    set_manual_unidade(uid, cod, valor)
    return jsonify({"ok": True})


@app.route("/api/parametros", methods=["POST"])
@login_required
def api_parametros():
    data = request.get_json(force=True)
    uid = resolve_unidade(data.get("unidade"))
    cod = str(data.get("codigo") or "").strip()
    if not cod:
        return jsonify({"ok": False, "erro": "Código inválido."}), 400
    try:
        est_min = float(data.get("estoque_minimo"))
        ponto = float(data.get("ponto_pedido"))
        caixa = float(data.get("caixa_com"))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "erro": "Valores inválidos."}), 400
    if est_min < 0 or ponto < 0 or caixa <= 0:
        return jsonify(
            {"ok": False, "erro": "Est. mínimo e ponto ≥ 0; caixa deve ser > 0."}
        ), 400
    try:
        save_catalogo_parametro(
            uid,
            cod,
            estoque_minimo=est_min,
            ponto_pedido=ponto,
            caixa_com=caixa,
            sqlite_path=_db_path(),
        )
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500
    return jsonify({"ok": True})


@app.route("/manual/extra", methods=["POST"])
@login_required
def manual_extra():
    descricao = (request.form.get("descricao") or "").strip()
    aba = (request.form.get("aba") or "MANUAL").strip() or "MANUAL"
    qtd_raw = request.form.get("quantidade") or ""
    cod_raw = (request.form.get("codigo") or "").strip()
    try:
        quantidade = float(qtd_raw)
    except ValueError:
        flash("Quantidade inválida.", "erro")
        return redirect(url_for("index"))
    if not descricao or quantidade <= 0:
        flash("Informe descrição e quantidade do item manual.", "erro")
        return redirect(url_for("index"))
    codigo = None
    if cod_raw:
        try:
            codigo = int(float(cod_raw))
        except ValueError:
            flash("Código inválido.", "erro")
            return redirect(url_for("index"))
    extras = itens_extras_manuais()
    extras.append(
        {
            "codigo": codigo,
            "descricao": descricao,
            "aba": aba,
            "quantidade": quantidade,
        }
    )
    session["extras_manuais"] = extras
    session.modified = True
    flash("Item manual adicionado ao pedido.", "ok")
    return redirect(url_for("index", filtro="pedir"))


@app.route("/manual/extra/<int:idx>/remover", methods=["POST"])
@login_required
def manual_extra_remover(idx: int):
    extras = itens_extras_manuais()
    if 0 <= idx < len(extras):
        extras.pop(idx)
        session["extras_manuais"] = extras
        session.modified = True
        flash("Item manual removido.", "ok")
    return redirect(url_for("index"))


@app.route("/pedido/salvar", methods=["POST"])
@login_required
def pedido_salvar():
    uid = resolve_unidade()
    data_ped = parse_data_pedido(request.form.get("data_pedido"))
    observacao = (request.form.get("observacao") or "").strip()
    itens = coletar_itens_pedido(incluir_extras=True, unidade=uid)
    if not itens:
        flash("Não há itens para salvar (Pedir = Sim ou manuais).", "erro")
        return redirect(url_for("index", unidade=uid))
    try:
        user = current_user() or {}
        # prefixa observação com a unidade para distinguir no histórico
        obs = observacao
        tag = f"[{unidade_cfg(uid)['titulo']}]"
        if not (obs or "").startswith(tag):
            obs = f"{tag} {obs}".strip()
        pid = criar_pedido(
            data_pedido=data_ped,
            usuario=user.get("nome") or user.get("username"),
            observacao=obs,
            itens=itens,
            sqlite_path=_db_path(),
            unidade=uid,
        )
        session["extras_manuais"] = []
        session.modified = True
        flash(f"Pedido #{pid} salvo ({data_ped.strftime('%d/%m/%Y')}) — {len(itens)} itens.", "ok")
        return redirect(url_for("historico_detalhe", pedido_id=pid))
    except Exception as e:
        flash(f"Erro ao salvar pedido: {e}", "erro")
        return redirect(url_for("index", unidade=uid))


@app.route("/pedido/pdf")
@login_required
def pedido_pdf_atual():
    uid = resolve_unidade()
    cfg = unidade_cfg(uid)
    data_ped = parse_data_pedido(request.args.get("data"))
    aba = request.args.get("aba") or "todos"
    filtro = request.args.get("filtro") or "todos"
    q = request.args.get("q") or ""

    itens_view, meta = listar_itens_filtrados(aba, filtro, q, unidade=uid)
    itens = itens_para_pdf(itens_view)
    if not itens:
        flash("Nenhum item no filtro atual para gerar PDF.", "erro")
        return redirect(url_for("index", unidade=uid, aba=aba, filtro=filtro, q=q))

    nome = cfg["titulo"]
    if meta["ver_todos"]:
        titulo = f"Pedido {nome} — visão atual (todas as categorias)"
    else:
        titulo = f"Pedido {nome} — {itens[0].get('aba') or aba}"
    if filtro == "pedir":
        titulo += " · só a pedir"
    elif filtro == "sem":
        titulo += " · sem estoque"
    if q:
        titulo += f" · busca: {q}"

    user = current_user() or {}
    pdf = gerar_pdf_pedido(
        titulo=titulo,
        data_pedido=data_ped.strftime("%d/%m/%Y"),
        usuario=user.get("nome") or user.get("username"),
        itens=itens,
    )
    return send_file(
        io.BytesIO(pdf),
        as_attachment=True,
        download_name=f"pedido_{uid}_{data_ped.isoformat()}.pdf",
        mimetype="application/pdf",
    )


@app.route("/historico")
@login_required
def historico():
    f_unidade = (request.args.get("unidade") or "").strip() or None
    f_aba = (request.args.get("aba") or "").strip() or None
    f_usuario = (request.args.get("usuario") or "").strip() or None
    if f_unidade in ("centro", "centro_cirurgico", "bloco"):
        f_unidade = "cc"
    try:
        # garante coluna unidade
        init_db(_db_path())
        pedidos = listar_pedidos(
            _db_path(), unidade=f_unidade, aba=f_aba, usuario=f_usuario
        )
        filtros = listar_filtros_historico(_db_path())
    except Exception as e:
        pedidos = []
        filtros = {"usuarios": [], "abas": [], "unidades": []}
        flash(f"Erro ao carregar histórico: {e}", "erro")

    # títulos de unidade conhecidos + ids vindos do banco
    unidade_opcoes = []
    seen = set()
    for uid, cfg in UNIDADES.items():
        unidade_opcoes.append({"id": uid, "titulo": cfg["titulo"]})
        seen.add(uid)
    for uid in filtros.get("unidades") or []:
        if uid and uid not in seen:
            unidade_opcoes.append({"id": uid, "titulo": uid})
            seen.add(uid)

    # agrupar para exibição: unidade → usuário → pedidos
    grupos = []
    atual_u = object()
    atual_user = object()
    bloco_u = None
    bloco_user = None
    for p in pedidos:
        uid = p.get("unidade") or "ufcd"
        user_name = p.get("usuario") or "—"
        if uid != atual_u:
            atual_u = uid
            atual_user = object()
            bloco_u = {
                "unidade_id": uid,
                "unidade_titulo": UNIDADES.get(uid, {}).get("titulo") or uid,
                "usuarios": [],
            }
            grupos.append(bloco_u)
        if user_name != atual_user:
            atual_user = user_name
            bloco_user = {"usuario": user_name, "pedidos": []}
            bloco_u["usuarios"].append(bloco_user)
        bloco_user["pedidos"].append(p)

    return render_template(
        "historico.html",
        pedidos=pedidos,
        grupos=grupos,
        filtros=filtros,
        unidade_opcoes=unidade_opcoes,
        f_unidade=f_unidade or "",
        f_aba=f_aba or "",
        f_usuario=f_usuario or "",
        user=current_user(),
        is_admin=is_admin(),
        usando_neon=using_neon(),
    )


@app.route("/historico/<int:pedido_id>")
@login_required
def historico_detalhe(pedido_id: int):
    ped = obter_pedido(pedido_id, _db_path())
    if not ped:
        flash("Pedido não encontrado.", "erro")
        return redirect(url_for("historico"))
    uid = ped.get("unidade") or "ufcd"
    # agrupa itens por categoria (subitem)
    por_aba = []
    atual = None
    bloco = None
    for it in ped.get("itens") or []:
        aba = it.get("aba") or "Sem categoria"
        if aba != atual:
            atual = aba
            bloco = {"aba": aba, "itens": []}
            por_aba.append(bloco)
        bloco["itens"].append(it)
    return render_template(
        "historico_detalhe.html",
        pedido=ped,
        por_aba=por_aba,
        unidade_titulo=UNIDADES.get(uid, {}).get("titulo") or uid,
        user=current_user(),
        is_admin=is_admin(),
    )


@app.route("/historico/<int:pedido_id>/pdf")
@login_required
def historico_pdf(pedido_id: int):
    ped = obter_pedido(pedido_id, _db_path())
    if not ped:
        flash("Pedido não encontrado.", "erro")
        return redirect(url_for("historico"))
    data_ped = ped["data_pedido"]
    if hasattr(data_ped, "strftime"):
        data_str = data_ped.strftime("%d/%m/%Y")
        data_iso = data_ped.isoformat()
    else:
        data_str = str(data_ped)
        data_iso = str(data_ped)
    pdf = gerar_pdf_pedido(
        titulo=f"Pedido #{pedido_id} — {UNIDADES.get(ped.get('unidade') or 'ufcd', {}).get('titulo') or ped.get('unidade') or 'UFCD'}",
        data_pedido=data_str,
        usuario=ped.get("usuario"),
        itens=ped.get("itens") or [],
        observacao=ped.get("observacao"),
    )
    return send_file(
        io.BytesIO(pdf),
        as_attachment=True,
        download_name=f"pedido_{pedido_id}_{data_iso}.pdf",
        mimetype="application/pdf",
    )


@app.route("/historico/<int:pedido_id>/excluir", methods=["POST"])
@admin_required
def historico_excluir(pedido_id: int):
    if excluir_pedido(pedido_id, _db_path()):
        flash("Pedido excluído.", "ok")
    else:
        flash("Pedido não encontrado.", "erro")
    return redirect(url_for("historico"))


@app.route("/exportar")
@login_required
def exportar():
    uid = resolve_unidade()
    cfg = unidade_cfg(uid)
    data_ped = parse_data_pedido(request.args.get("data"))
    aba = request.args.get("aba") or "todos"
    filtro = request.args.get("filtro") or "todos"
    q = request.args.get("q") or ""
    itens, meta = listar_itens_filtrados(aba, filtro, q, unidade=uid)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pedido"

    header_fill = PatternFill("solid", fgColor="16352B")
    header_font = Font(color="FFFFFF", bold=True)
    sim_fill = PatternFill("solid", fgColor="FDEBD0")
    estoque_fill = PatternFill("solid", fgColor="EAF5FB")
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    ws["A1"] = f"PEDIDO {cfg['titulo'].upper()} — DATA: {data_ped.strftime('%d/%m/%Y')}"
    ws["A2"] = (
        f"Filtro: aba={meta['aba']} · {meta['filtro']}"
        + (f" · busca={meta['q']}" if meta["q"] else "")
    )
    if cfg.get("mostra_ponto_caixa", True):
        headers = [
            "Categoria",
            "Cód. AGHU",
            "Medicamento",
            "Est. Mínimo",
            "Ponto de Pedido",
            "Caixa com",
            "Estoque",
            "Pedir?",
            "Quanto Pedir?",
        ]
    else:
        headers = [
            "Categoria",
            "Cód. AGHU",
            "Item",
            "Est. Mínimo",
            "Estoque",
            "Pedir?",
            "Quanto Pedir?",
        ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(4, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin

    for r, it in enumerate(itens, 5):
        if cfg.get("mostra_ponto_caixa", True):
            vals = [
                it.get("aba") or "",
                it["codigo"],
                it["descricao"],
                it["estoque_minimo"],
                it["ponto_pedido"],
                it["caixa_com"],
                "" if it["estoque_aghu"] is None else it["estoque_aghu"],
                "Sim" if it["pedir"] else "Não",
                "" if it["quanto_pedir"] is None else it["quanto_pedir"],
            ]
            estoque_col = 7
            qtde_col = 9
        else:
            vals = [
                it.get("aba") or "",
                it["codigo"],
                it["descricao"],
                it["estoque_minimo"],
                "" if it["estoque_aghu"] is None else it["estoque_aghu"],
                "Sim" if it["pedir"] else "Não",
                "" if it["quanto_pedir"] is None else it["quanto_pedir"],
            ]
            estoque_col = 5
            qtde_col = 7
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.border = thin
            if it["pedir"] and c == qtde_col:
                cell.fill = sim_fill
                cell.font = Font(bold=True, color="8A3B00")
            elif c == estoque_col and not it["pedir"]:
                cell.fill = estoque_fill
            elif c == estoque_col:
                cell.fill = estoque_fill

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 42
    for col in "DEFGHI":
        ws.column_dimensions[col].width = 14

    if not itens:
        ws["A5"] = "Nenhum item no filtro atual."

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"pedido_{uid}_{data_ped.isoformat()}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/status")
@login_required
def api_status():
    cat = load_catalogo()
    return jsonify(
        {
            "abas": len(cat.get("abas", [])),
            "itens": sum(a.get("count", 0) for a in cat.get("abas", [])),
            "estoque": len(estoque_saldos()),
            "meta": estoque_meta(),
        }
    )


def bootstrap_estoque_exemplo():
    """Se existir exemplo e ainda não houver estoque, importa."""
    exemplo = DATA / "EstoqueFarmacia-exemplo.xlsx"
    if exemplo.exists() and not ESTOQUE_PATH.exists():
        mapa, meta = parse_estoque_xlsx(exemplo)
        save_estoque(mapa, meta)


# Carrega estoque de exemplo na subida (local e Vercel)
try:
    bootstrap_estoque_exemplo()
except Exception:
    pass


if __name__ == "__main__":
    print("Pedidos UFCD — http://127.0.0.1:5000")
    app.run(debug=not ON_VERCEL, port=int(os.environ.get("PORT", 5000)))
